"""Smoke tests des scripts d'export de métadonnées (`tools/`).

`tools/` est hors couverture, mais ces tests verrouillent la **non-régression
fonctionnelle** des quatre outils : chacun tourne de bout en bout sur un corpus
réel (bâti par l'API) et produit une sortie exploitable. Exécution en
SOUS-PROCESSUS (comme le test `live`), avec `BD_DB_PATH` pointant la base de test.
"""
import json
import os
import re
import sqlite3
import subprocess
import sys
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
TOOLS = REPO_ROOT / "tools"


def _run(script, db_path, data_dir, *args):
    # Les tools forcent EUX-MÊMES leur stdout/stderr en UTF-8 (_commun.forcer_utf8), donc on
    # décode en UTF-8 SANS PYTHONUTF8. Ce test EXERCE ainsi le garde de portabilité Windows :
    # sans lui, un tool crasherait ici sur un caractère hors cp1252 (emoji d'un résumé PyPI,
    # « → »…). Cf. tools/_commun.forcer_utf8.
    env = {**os.environ, "BD_DB_PATH": str(db_path), "BD_DATA_DIR": str(data_dir)}
    return subprocess.run([sys.executable, str(TOOLS / script), *args],
                          cwd=str(REPO_ROOT), env=env, capture_output=True,
                          text=True, encoding="utf-8")


@pytest.fixture
def corpus(client, region, db_path, data_dir):
    """Album + planche + région (via l'API), plus une annotation dont la note est une
    FORMULE Excel (pour le garde anti-injection du XLSX). Checkpoint WAL pour qu'un
    process lecteur séparé voie tout."""
    conn = sqlite3.connect(db_path)
    conn.execute("INSERT INTO annotations (region_id, note) VALUES (?, ?)",
                 (region["id"], "=1+1"))
    conn.commit()
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.close()
    return {"db": db_path, "data": data_dir}


def test_metadonnees_json(corpus):
    r = _run("metadonnees_collection.py", corpus["db"], corpus["data"], "--json", "-")
    assert r.returncode == 0, r.stderr
    mc = json.loads(r.stdout)["metadonnees_collection"]
    assert len(mc["albums"]) == 1
    assert mc["paradonnee"]["schema_version"]
    assert mc["paradonnee"]["outil"]["nom"] == "BéDéditeur"   # parité outil


def test_description_json(corpus):
    r = _run("description_collection.py", corpus["db"], corpus["data"], "--json", "-")
    assert r.returncode == 0, r.stderr
    doc = json.loads(r.stdout)["description_collection"]
    assert doc["couverture"]["albums"] == 1
    assert doc["outil"]["nom"] == "BéDéditeur"                # même provenance que les enreg.


def test_csv_tables_bom(corpus, tmp_path):
    dossier = tmp_path / "tables"
    r = _run("metadonnees_collection.py", corpus["db"], corpus["data"],
             "--csv-dir", str(dossier))
    assert r.returncode == 0, r.stderr
    assert (dossier / "albums.csv").read_bytes()[:3] == b"\xef\xbb\xbf"   # BOM UTF-8
    assert (dossier / "paradonnee.csv").exists()


def test_xlsx_anti_injection(corpus, tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    xlsx = tmp_path / "meta.xlsx"
    r = _run("metadonnees_collection.py", corpus["db"], corpus["data"], "--xlsx", str(xlsx))
    assert r.returncode == 0, r.stderr
    wb = load_workbook(xlsx)
    assert {"fiche", "arbre", "paradonnee"} <= set(wb.sheetnames)
    # la note '=1+1' doit être du TEXTE, jamais une formule
    cells = [c for row in wb["annotations"].iter_rows() for c in row if c.value == "=1+1"]
    assert cells and all(c.data_type == "s" for c in cells)


def test_iiif_valide(corpus, tmp_path):
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out))
    assert r.returncode == 0, r.stderr
    assert (out / "collection.json").exists()
    v = _run("valider_iiif.py", corpus["db"], corpus["data"], str(out))
    assert v.returncode == 0, v.stdout + v.stderr   # manifests conformes


def test_collection_crud_et_portee(corpus, album, client, db_path, tmp_path):
    """Palier COLLECTION (v14) : l'outil `gerer_collections.py` crée une collection et y
    range un album ; `--collection` restreint alors les exports (records + fiche + IIIF) à
    ses albums, l'identité étant renseignée depuis la ligne `collection`. On sème un SECOND
    album (hors sélection) pour prouver la portée."""
    autre = client.post("/api/albums", json={"titre": "Hors sélection"}).json()
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")   # le lecteur RO (sous-process) voit tout
    conn.close()

    creer = _run("gerer_collections.py", corpus["db"], corpus["data"], "creer",
                 "--nom", "Sélection", "--licence", "CC-BY-4.0", "--statut", "public",
                 "--responsable", "Hugo;chercheur", "--albums", str(album["id"]))
    assert creer.returncode == 0, creer.stderr
    cid = creer.stdout.strip()
    assert cid.isdigit(), f"l'id de collection attendu sur stdout, reçu : {creer.stdout!r}"

    # Records scopés : seul l'album de la collection, bloc `collection` renseigné.
    m = _run("metadonnees_collection.py", corpus["db"], corpus["data"],
             "--json", "-", "--collection", cid)
    assert m.returncode == 0, m.stderr
    doc = json.loads(m.stdout)["metadonnees_collection"]
    ids = [a["id"] for a in doc["albums"]]
    assert ids == [album["id"]] and autre["id"] not in ids
    assert doc["perimetre"]["collection_id"] == int(cid)
    assert doc["collection"]["nom"] == "Sélection"
    assert doc["collection"]["responsables"][0]["nom"] == "Hugo"

    # Fiche scopée : identité depuis la ligne collection, couverture restreinte (1/2 albums).
    d = _run("description_collection.py", corpus["db"], corpus["data"],
             "--json", "-", "--collection", cid)
    assert d.returncode == 0, d.stderr
    fiche = json.loads(d.stdout)["description_collection"]
    assert fiche["identite"]["nom"] == "Sélection"
    assert fiche["identite"]["statut_diffusion"] == "public"
    assert fiche["couverture"]["albums"] == 1

    # IIIF scopé : la Collection prend le nom du jeu, un seul manifest.
    out = tmp_path / "iiif"
    v = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out), "--collection", cid)
    assert v.returncode == 0, v.stderr
    coll = json.loads((out / "collection.json").read_text(encoding="utf-8"))
    assert coll["label"]["fr"] == ["Sélection"]
    assert len(coll["items"]) == 1


def test_contributions_et_edition_dans_export(corpus, album, client, db_path):
    """N0 (v15) : contributions (nom+rôle résolu) et champs d'édition apparaissent dans les
    records exportés (arbre JSON)."""
    client.put(f"/api/albums/{album['id']}", json={"date_edition": "1960", "langue": "fr"})
    client.post(f"/api/albums/{album['id']}/contributions",
                json={"nom": "Hergé", "role": "scénariste"})
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.close()
    r = _run("metadonnees_collection.py", corpus["db"], corpus["data"], "--json", "-")
    assert r.returncode == 0, r.stderr
    doc = json.loads(r.stdout)["metadonnees_collection"]
    alb = doc["albums"][0]
    assert alb["date_edition"] == "1960" and alb["langue"] == "fr"
    assert alb["contributions"] == [
        {"nom": "Hergé", "role": "scénariste", "bucket": "creator", "marc": "aut"}]
    assert any(role["label"] == "scénariste" for role in doc["contribution_roles"])


def test_qualite_dans_fiche(corpus):
    """Paradonnée QUALITÉ : la fiche porte un bloc `qualite` (relecture + accords),
    `accord_inter` étant marqué de portée corpus (le journal n'est pas re-scopé)."""
    r = _run("description_collection.py", corpus["db"], corpus["data"], "--json", "-")
    assert r.returncode == 0, r.stderr
    q = json.loads(r.stdout)["description_collection"]["qualite"]
    assert set(q) == {"relecture", "accord_modele", "accord_inter"}
    assert set(q["relecture"]) >= {"a_faire", "en_cours", "faite", "pct_faite"}
    assert q["accord_inter"]["portee"] == "corpus"


def test_relecture_statut_dans_records(corpus, tmp_path):
    """B5 : le statut de relecture (dérivé) sort sur les planches — arbre JSON + table CSV.
    Sans token relu, une planche est 'a_faire'."""
    r = _run("metadonnees_collection.py", corpus["db"], corpus["data"], "--json", "-")
    assert r.returncode == 0, r.stderr
    pl = json.loads(r.stdout)["metadonnees_collection"]["albums"][0]["planches"][0]
    assert pl["relecture_statut"] == "a_faire"
    dossier = tmp_path / "t"
    assert _run("metadonnees_collection.py", corpus["db"], corpus["data"],
                "--csv-dir", str(dossier)).returncode == 0
    entete = (dossier / "planches.csv").read_text(encoding="utf-8-sig").splitlines()[0]
    assert "relecture_statut" in entete.split(",")


def test_qualite_onglet_xlsx(corpus, tmp_path):
    """XLSX : onglet `qualite` (tableau de bord) + colonne relecture_statut sur `planches`."""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    xlsx = tmp_path / "m.xlsx"
    r = _run("metadonnees_collection.py", corpus["db"], corpus["data"], "--xlsx", str(xlsx))
    assert r.returncode == 0, r.stderr
    wb = load_workbook(xlsx)
    assert "qualite" in wb.sheetnames
    sections = {row[0].value for row in wb["qualite"].iter_rows(min_row=2)}
    assert {"relecture", "accord modèle", "accord inter"} <= sections
    assert "relecture_statut" in [c.value for c in wb["planches"][1]]


def test_accord_modele_scope_par_collection(client, db_path, data_dir):
    """L'accord modèle↔humain de la fiche est SCOPÉ à la collection (param `album_ids` du
    cœur `accord.rapport`). On sème deux albums relus (un dans la collection, un hors) et on
    vérifie que la fiche scopée ne compte que l'échantillon relu du périmètre."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys=ON")

    def semer(titre, lemme_corr):
        aid = conn.execute("INSERT INTO albums(titre) VALUES(?)", (titre,)).lastrowid
        pid = conn.execute("INSERT INTO planches(album_id, numero, chemin_web) "
                           "VALUES(?,1,'x.jpg')", (aid,)).lastrowid
        rid = conn.execute("INSERT INTO regions(planche_id, type, ordre) "
                           "VALUES(?, 'bulle', 1)", (pid,)).lastrowid
        conn.execute("INSERT INTO tokens(region_id, ordre, texte, lemme, pos, morph) "
                     "VALUES(?,0,'m','chat','NOUN','')", (rid,))
        conn.execute("INSERT INTO token_correction(region_id, ordre, forme, lemme, pos, "
                     "morph, etat, obsolete) VALUES(?,0,'m',?, 'NOUN','', 'corrige', 0)",
                     (rid, lemme_corr))
        return aid

    a_in = semer("Dans", "chat")       # correction lemme == auto → accord sur le lemme
    semer("Hors", "chien")             # correction lemme != auto → désaccord (hors périmètre)
    conn.commit()
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.close()

    creer = _run("gerer_collections.py", db_path, data_dir, "creer",
                 "--nom", "Q", "--albums", str(a_in))
    assert creer.returncode == 0, creer.stderr
    cid = creer.stdout.strip()

    def fiche(*extra):
        r = _run("description_collection.py", db_path, data_dir, "--json", "-", *extra)
        assert r.returncode == 0, r.stderr
        return json.loads(r.stdout)["description_collection"]["qualite"]["accord_modele"]

    assert fiche()["revus"] == 2                          # corpus entier : 2 relus
    scoped = fiche("--collection", cid)
    assert scoped["revus"] == 1                           # collection : 1 relu
    assert scoped["champs"]["lemme"]["accord"] == 1       # ce relu accorde sur le lemme


def test_iiif_conformance_stricte(corpus, tmp_path):
    """Conformité STRICTE via iiif-prezi3 (lib IIIF officielle) : le manifest généré se
    re-parse sans erreur dans ses modèles typés → validation INDÉPENDANTE de notre
    validateur maison. Skip propre si la lib n'est pas installée."""
    pytest.importorskip("iiif_prezi3")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out))
    assert r.returncode == 0, r.stderr
    v = _run("valider_iiif.py", corpus["db"], corpus["data"], str(out))
    assert v.returncode == 0, v.stdout + v.stderr
    assert "Conformité stricte (iiif-prezi3) : exécutée" in v.stdout   # la passe a bien tourné


# --------------------------------------------------------------------------- #
# DROIT-1 — publier n'est pas citer
#
# Le manifeste IIIF est le SEUL artefact du dépôt qui émette des URL d'images vers
# l'extérieur : c'est donc le point où le régime de diffusion devient opposable. La règle
# est fail-closed et tient en une phrase — publier suppose de NOMMER la collection qu'on
# publie. Elle règle du même coup le cas d'un album vivant dans plusieurs collections
# (AUTH-3) sans inventer d'arbitrage.
# --------------------------------------------------------------------------- #
def _images_du_manifeste(chemin):
    """Les URL d'images peintes sur les Canvas d'un manifeste."""
    man = json.loads(chemin.read_text(encoding="utf-8"))
    return [it["body"]["id"] for c in man["items"] for page in c["items"]
            for it in page["items"] if it.get("motivation") == "painting"]


def _collection(corpus, album_id, nom, statut=None, date_embargo=None):
    args = ["creer", "--nom", nom, "--albums", str(album_id)]
    if statut:
        args += ["--statut", statut]
    if date_embargo:
        args += ["--date-embargo", date_embargo]
    r = _run("gerer_collections.py", corpus["db"], corpus["data"], *args)
    assert r.returncode == 0, r.stderr
    return r.stdout.strip()


def _declaration(chemin):
    """Le texte du `requiredStatement` d'un manifeste, ou "" s'il n'y en a pas."""
    man = json.loads(chemin.read_text(encoding="utf-8"))
    val = man.get("requiredStatement", {}).get("value", {})
    return " ".join(v for lst in val.values() for v in lst)


def test_iiif_sans_collection_nommee_n_emporte_pas_d_images(corpus, album, tmp_path):
    """Sans `--collection`, l'outil porte sur le corpus entier — donc sur AUCUN régime
    déclaré. Il écrit alors un manifeste sans images, et le dit.

    Fail-closed : l'absence de déclaration ne vaut pas autorisation. C'est aussi ce qui
    évite d'avoir à arbitrer entre les régimes des collections d'un album partagé.
    """
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out))
    assert r.returncode == 0, r.stderr
    assert "SANS IMAGES" in r.stderr
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json") == []


def test_iiif_d_une_collection_publique_emporte_les_images(corpus, album, tmp_path):
    """Le pendant : une collection déclarée `public` publie ses scans. Sans quoi le
    chantier n'aurait fait que casser la publication légitime."""
    cid = _collection(corpus, album["id"], "Domaine public", "public")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "SANS IMAGES" not in r.stderr
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json")


def test_iiif_d_une_collection_restreinte_publie_sans_les_scans(corpus, album, tmp_path):
    """Le scénario de la piste A : déposer ouvertement son enrichissement sur un fonds
    qu'on ne peut pas diffuser. Le Canvas SURVIT sans image — il garde ses dimensions et
    ses annotations de régions, donc la géométrie et le travail restent publiables."""
    cid = _collection(corpus, album["id"], "Sous droits", "restreint")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "restreint" in r.stderr and "SANS IMAGES" in r.stderr
    man = json.loads((out / f"manifest-a{album['id']}.json").read_text(encoding="utf-8"))
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json") == []
    assert man["items"], "les Canvas restent : la géométrie est publiable"
    assert man["items"][0]["height"] and man["items"][0]["width"]


def test_iiif_verbatim_refuse_hors_public(corpus, album, tmp_path):
    """`--verbatim` fait sortir le TEXTE de l'œuvre. Publier le texte d'un fonds sous
    droits est de la diffusion, pas de la citation : l'outil refuse, et renvoie vers le
    geste qui convient (l'export de figure)."""
    cid = _collection(corpus, album["id"], "Sous droits", "restreint")
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(tmp_path / "x"),
             "--collection", cid, "--verbatim")
    assert r.returncode == 2
    assert "REFUS" in r.stderr and "citez plutôt" in r.stderr


def test_l_exemption_tient_a_la_declaration_pas_a_l_absence(tmp_path):
    """Le garde-fou du garde-fou.

    Un Canvas sans peinture est légitime quand le manifeste DÉCLARE retenir ses scans
    (`requiredStatement`). Sans cette condition, l'exemption excuserait aussi un manifeste
    qui a simplement OUBLIÉ ses images — et la règle ne mesurerait plus rien.
    """
    sys.path.insert(0, str(REPO_ROOT / "tools"))
    import valider_iiif as v
    nu = {"@context": "http://iiif.io/api/presentation/3/context.json",
          "id": "http://x/manifest.json", "type": "Manifest",
          "label": {"fr": ["A"]},
          "items": [{"id": "http://x/c1", "type": "Canvas", "label": {"fr": ["p1"]},
                     "height": 100, "width": 80, "items": []}]}
    assert v.valider_manifest(nu).err, "un oubli d'images doit rester une erreur"
    declare = dict(nu, requiredStatement={
        "label": {"fr": [v.DECLARATION_SANS_IMAGES]},
        "value": {"fr": ["Les images de ce corpus ne sont pas diffusées."]}})
    assert not v.valider_manifest(declare).err
    assert v.valider_manifest(declare).warn      # dit quand même, en avertissement


# --------------------------------------------------------------------------- #
# DROIT-1 — la date d'embargo RETIENT, elle ne PROMEUT jamais
# --------------------------------------------------------------------------- #
# Décidé le 2026-08-28, après avoir constaté que `_regime` ne lisait que `statut_diffusion`
# et publiait donc les scans d'une collection déclarée `public` dont l'embargo courait
# encore. L'asymétrie tient à ce que l'outil IGNORE : il ne sait pas POURQUOI l'embargo
# existe. Un délai qu'on s'est donné se lève de soi-même ; un délai imposé par un ayant
# droit ne se lève pas — son échéance dit que la contrainte a couru, pas que les droits
# sont à nous. Publier sur la foi d'une date serait une politique inventée (DEPOT-1).
# --------------------------------------------------------------------------- #
def test_embargo_pendant_retient_les_scans_d_une_collection_publique(corpus, album, tmp_path):
    """Le trou trouvé en relecture : `public` + embargo qui court publiait quand même.

    La date est plus restrictive que le statut, donc la date gagne — fail-closed, dans la
    seule direction qui ne peut pas nuire.
    """
    cid = _collection(corpus, album["id"], "Publique plus tard", "public", "2099-01-01")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "SANS IMAGES" in r.stderr and "2099-01-01" in r.stderr
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json") == []


def test_embargo_pendant_le_manifeste_ne_dit_pas_qu_il_est_public(corpus, album, tmp_path):
    """Et il le DIT correctement. Un manifeste amputé qui annonce « régime : public » se
    contredit lui-même, et le lecteur n'a aucun moyen de savoir quelle moitié croire —
    c'est le message d'erreur qui ment, déjà attrapé une fois sur AUTH-3."""
    cid = _collection(corpus, album["id"], "Publique plus tard", "public", "2099-01-01")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    decl = _declaration(out / f"manifest-a{album['id']}.json")
    assert "embargo" in decl and "2099-01-01" in decl, decl


def test_embargo_echu_ne_publie_pas_tout_seul(corpus, album, tmp_path):
    """Le cœur de l'arbitrage : la date ne PROMEUT jamais. Une collection `embargo` dont
    l'échéance est passée reste non publiée — la passer en `public` est un acte, avec
    quelqu'un derrière."""
    cid = _collection(corpus, album["id"], "Embargo fini", "embargo", "2020-01-01")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json") == []


def test_embargo_echu_cesse_d_etre_muet(corpus, album, tmp_path):
    """Ne rien faire ne veut pas dire se taire. Un embargo échu que personne ne remarque
    garde un corpus fermé par INERTIE, ce qui trahit l'orientation open-science aussi
    sûrement qu'une fuite trahit les droits : l'outil le signale là où quelqu'un s'apprête
    justement à publier."""
    cid = _collection(corpus, album["id"], "Embargo fini", "embargo", "2020-01-01")
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(tmp_path / "iiif"),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "ÉCHU" in r.stderr and "2020-01-01" in r.stderr


def test_embargo_echu_et_declaree_publique_publie(corpus, album, tmp_path):
    """L'état final cohérent : l'échéance est passée ET quelqu'un a déclaré la collection
    `public`. Sans ce test, « retenir » et « ne jamais publier » seraient indistinguables."""
    cid = _collection(corpus, album["id"], "Embargo levé", "public", "2020-01-01")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "SANS IMAGES" not in r.stderr
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json")


def test_date_d_embargo_illisible_retient_et_le_dit(corpus, album, tmp_path):
    """`date_embargo` est du texte libre, et rien ne l'impose à l'écriture. Une date qu'on
    ne sait pas lire RETIENT — sinon une faute de frappe ouvrirait la porte — et elle se
    dit, sinon la faute de frappe passerait pour une décision."""
    cid = _collection(corpus, album["id"], "Date fautive", "public", "bientôt")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "illisible" in r.stderr and "AAAA-MM-JJ" in r.stderr
    assert _images_du_manifeste(out / f"manifest-a{album['id']}.json") == []


def test_verbatim_refuse_sous_embargo_pendant(corpus, album, tmp_path):
    """`--verbatim` fait sortir le TEXTE de l'œuvre. La garde suit la même règle que les
    images : une collection `public` sous embargo qui court n'est pas publiable."""
    cid = _collection(corpus, album["id"], "Publique plus tard", "public", "2099-01-01")
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(tmp_path / "x"),
             "--collection", cid, "--verbatim")
    assert r.returncode == 2, r.stderr
    assert "REFUS" in r.stderr and "2099-01-01" in r.stderr


def test_embargo_pendant_ne_promet_pas_ce_qui_n_arrivera_pas(corpus, album, tmp_path):
    """Relecture du 2026-08-28, sur du code écrit le jour même : le message annonçait
    « les images sortiront d'elles-mêmes une fois la date passée » pour TOUT embargo qui
    court. C'est vrai d'une collection déclarée `public`, et faux d'une `embargo` — à
    l'échéance elle reste `embargo`, donc non publiée. Le message qui ment, encore lui."""
    cid = _collection(corpus, album["id"], "Sous embargo", "embargo", "2099-01-01")
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(tmp_path / "iiif"),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "sortiront d'elles-mêmes" not in r.stderr, r.stderr
    assert "déclarer `public`" in r.stderr


def test_embargo_pendant_sur_publique_promet_bien_la_sortie(corpus, album, tmp_path):
    """Le pendant : là, la promesse est vraie — la date passée suffira."""
    cid = _collection(corpus, album["id"], "Publique plus tard", "public", "2099-01-01")
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(tmp_path / "iiif"),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    assert "sortiront d'elles-mêmes" in r.stderr


# --------------------------------------------------------------------------- #
# Nakala est l'entrepôt du FIGÉ — un artefact déposé doit dire de quand il date
# --------------------------------------------------------------------------- #
# Précision du 2026-08-28 : ShareDocs est le stockage VIVANT (modifiable, appelable à tout
# moment), Nakala l'entrepôt de ce qui est traité et FIGÉ. Le manifeste était le seul
# artefact de la chaîne de dépôt sans date — les notices posent `genere_le`, la figure
# citable `date_export`. Or c'est lui qu'on fige, et il porte une DÉCLARATION DE DROITS.
def _meta_manifeste(chemin, label):
    """La valeur d'une entrée `metadata` du manifeste, par son libellé."""
    man = json.loads(chemin.read_text(encoding="utf-8"))
    for m in man.get("metadata", []):
        if label in " ".join(v for lst in m["label"].values() for v in lst):
            return " ".join(v for lst in m["value"].values() for v in lst)
    return None


def test_le_manifeste_dit_de_quand_il_date(corpus, album, tmp_path):
    """Deux manifestes du même album déposés à un an d'intervalle seraient sinon
    indistinguables — et l'entrepôt ne les remplace pas, il les garde."""
    cid = _collection(corpus, album["id"], "Domaine public", "public")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    genere = _meta_manifeste(out / f"manifest-a{album['id']}.json", "généré le")
    assert genere and re.fullmatch(r"\d{4}-\d{2}-\d{2}", genere), genere


def test_la_declaration_de_droits_est_datee(corpus, album, tmp_path):
    """Une déclaration figée sans date se lit comme une vérité de toujours. Déposé,
    ce manifeste affirmera encore « régime : restreint » le jour où la collection sera
    passée `public` : datée, l'assertion redevient un constat vérifiable à la source."""
    cid = _collection(corpus, album["id"], "Sous droits", "restreint")
    out = tmp_path / "iiif"
    r = _run("iiif_manifest.py", corpus["db"], corpus["data"],
             "--base-url", "http://exemple/iiif", "--out-dir", str(out),
             "--collection", cid)
    assert r.returncode == 0, r.stderr
    chemin = out / f"manifest-a{album['id']}.json"
    genere = _meta_manifeste(chemin, "généré le")
    decl = _declaration(chemin)
    # La MÊME date des deux côtés : deux horodatages pour un seul export divergeraient.
    assert genere in decl, (genere, decl)
    assert "Constat du" in decl
