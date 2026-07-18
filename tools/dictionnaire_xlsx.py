"""Exporte le dictionnaire de métadonnées en classeur XLSX.

Convertit `docs/dictionnaire-metadonnees.md` en un classeur **une feuille par
section à tableau** (Conventions, paliers, Collection, Niveaux 0–8, Synthèse) +
un `_sommaire`. Régénérable à la demande → pas de binaire figé qui dériverait du
markdown. En-têtes gelés/gras, filtres, retour à la ligne, anti-injection de formule.

N'accède PAS à la base (pur markdown → xlsx). Requiert `openpyxl`.

Usage :
    python tools/dictionnaire_xlsx.py                    # → docs/exemples/dictionnaire-metadonnees.xlsx
    python tools/dictionnaire_xlsx.py --out fichier.xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
DICT = REPO / "docs" / "dictionnaire-metadonnees.md"
_INJECT = ("=", "+", "-", "@")
_SEP = re.compile(r"^\s*\|[\s:|-]+\|\s*$")     # ligne séparatrice d'un tableau markdown


def _nettoyer(cellule: str) -> str:
    """Retire le balisage markdown (gras/italique/code/liens) et les zéro-largeur."""
    c = cellule.replace("`", "")
    c = re.sub(r"\*\*(.+?)\*\*", r"\1", c)
    c = re.sub(r"\*(.+?)\*", r"\1", c)
    c = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", c)
    return c.replace("​", "").strip()


def _cellules(ligne: str) -> list[str]:
    """Découpe une ligne `| a | b |` en cellules, en respectant les pipes échappés `\\|`."""
    ligne = ligne.strip().replace(r"\|", "\x00")
    cells = [c.strip().replace("\x00", "|") for c in ligne.split("|")]
    if cells and cells[0] == "":
        cells = cells[1:]
    if cells and cells[-1] == "":
        cells = cells[:-1]
    return [_nettoyer(c) for c in cells]


def parser(md: str):
    """Renvoie [(titre_section, [(header, rows), ...])] — tableaux groupés par section `##`."""
    lignes = md.split("\n")
    ordre: list = []
    index: dict = {}
    titre = "Sans titre"
    i = 0
    while i < len(lignes):
        l = lignes[i]
        if l.startswith("## "):
            titre = _nettoyer(l[3:].strip())
            i += 1
            continue
        if l.lstrip().startswith("|") and i + 1 < len(lignes) and _SEP.match(lignes[i + 1]):
            header = _cellules(l)
            rows = []
            j = i + 2
            while j < len(lignes) and lignes[j].lstrip().startswith("|"):
                rows.append(_cellules(lignes[j]))
                j += 1
            if all(h == "" for h in header):        # légendes à en-tête vide
                header = ["clé", "sens"] if len(header) == 2 else \
                         [f"c{n + 1}" for n in range(len(header))]
            if titre not in index:
                index[titre] = len(ordre)
                ordre.append((titre, []))
            ordre[index[titre]][1].append((header, rows))
            i = j
            continue
        i += 1
    return ordre


def _nom_feuille(titre: str, pris: set) -> str:
    m = re.match(r"Niveau (\d+)\s*[—-]\s*(.+)", titre)
    if m:
        nom = f"N{m.group(1)} · {m.group(2)}"
    elif titre.startswith("Collection"):
        nom = "Collection"
    elif "palier" in titre.lower():
        nom = "Trois paliers"
    elif titre.startswith("Synthèse"):
        nom = "Synthèse"
    elif titre.startswith("Conventions"):
        nom = "Conventions"
    else:
        nom = titre
    nom = re.sub(r"[\[\]:*?/\\]", "-", nom)[:31].strip() or "Feuille"
    base, k = nom, 2
    while nom in pris:
        nom = f"{base[:28]} {k}"
        k += 1
    pris.add(nom)
    return nom


def ecrire(sections, chemin: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("XLSX demandé mais 'openpyxl' est absent "
                         "(pip install -r requirements-export.txt).")
    gras = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    wb = Workbook()
    wb.remove(wb.active)

    som = wb.create_sheet("_sommaire")
    som.append(["feuille", "section", "lignes"])

    pris: set = set()
    plan = [(_nom_feuille(titre, pris), titre, tables) for titre, tables in sections]

    for nom, titre, tables in plan:
        som.append([nom, titre, sum(len(rows) for _, rows in tables)])
    for cell in som[1]:
        cell.font = gras
    som.freeze_panes = "A2"
    som.column_dimensions["A"].width = 22
    som.column_dimensions["B"].width = 44

    for nom, titre, tables in plan:
        ws = wb.create_sheet(nom)
        ligne = 1
        for ti, (header, rows) in enumerate(tables):
            if ti > 0:
                ligne += 1                          # ligne vide entre tableaux empilés
            for c, h in enumerate(header, 1):
                cell = ws.cell(ligne, c, h)
                cell.font = gras
                cell.alignment = wrap
            for k, rr in enumerate(rows, 1):
                for c, v in enumerate(rr, 1):
                    cell = ws.cell(ligne + k, c, v)
                    cell.alignment = wrap
                    if isinstance(v, str) and v[:1] in _INJECT:
                        cell.data_type = "s"        # anti-injection de formule
            ligne += 1 + len(rows)

        largeurs: dict = {}
        for r in ws.iter_rows():
            for cell in r:
                if cell.value is not None:
                    largeurs[cell.column] = max(largeurs.get(cell.column, 0),
                                                min(len(str(cell.value)), 60))
        for col, w in largeurs.items():
            ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 12), 60)

        ws.freeze_panes = "A2"
        if len(tables) == 1:                        # filtre sur les feuilles à tableau unique
            header, rows = tables[0]
            ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"

    wb.save(chemin)


def main(argv=None) -> int:
    from _commun import forcer_utf8
    forcer_utf8()                             # Windows : stdout/stderr en UTF-8 (cp1252 sinon)
    ap = argparse.ArgumentParser(description="Dictionnaire de métadonnées → classeur XLSX.")
    ap.add_argument("--out", default=str(REPO / "docs" / "exemples"
                                        / "dictionnaire-metadonnees.xlsx"),
                    help="fichier de sortie (défaut : docs/exemples/dictionnaire-metadonnees.xlsx)")
    ap.add_argument("--source", default=str(DICT), help="markdown source")
    args = ap.parse_args(argv)

    md = Path(args.source).read_text(encoding="utf-8")
    sections = parser(md)
    if not sections:
        raise SystemExit(f"Aucun tableau trouvé dans {args.source}")
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    ecrire(sections, args.out)
    print(f"Classeur écrit : {args.out} ({len(sections)} feuilles + sommaire)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
