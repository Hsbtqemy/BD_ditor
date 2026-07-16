"""Export des MÉTADONNÉES d'une collection — les enregistrements eux-mêmes.

À NE PAS confondre avec `description_collection.py` (la *fiche* : quels champs on
emploie + leur couverture — méta-niveau). Ici on sort **les valeurs réelles**,
une entité à la fois, en respectant la hiérarchie du corpus :

    collection → albums → planches → régions (case ⊃ bulle) → tokens

Deux vues du MÊME modèle (les mêmes dérivations : citation, numéro éditorial,
locuteur/présence par nom, attributs, tokens effectifs) :

  • `--json`   : arbre imbriqué (une collection = un document) ;
  • `--csv-dir`/`--zip` : tables relationnelles à plat (un fichier par niveau,
                 recollables par les clés album_id/planche_id/region_id).

Les entités de niveau corpus (personnages, vocabulaire facetté) sont sorties une
fois et référencées depuis les régions. Lecture SEULE.

Le texte OCR est du CONTENU (`restreint`), pas de la métadonnée : par défaut on
n'expose que présence + longueur. `--verbatim` inclut le texte (export détenu).

Périmètre par défaut : le corpus entier (l'entité `collection` n'existe pas encore).

Usage :
    python tools/metadonnees_collection.py --json f.json
    python tools/metadonnees_collection.py --csv-dir tables/
    python tools/metadonnees_collection.py --zip metadonnees.zip
La base suit la config du projet (BD_DB_PATH / BD_DATA_DIR).
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sqlite3
import sys
import zipfile
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DB_PATH, BASE_DIR  # noqa: E402
import database  # noqa: E402  (réutilise numeros_editoriaux / citations_regions)
from _commun import version_outil, environnement, composants  # noqa: E402  (provenance / env, partagés)


def _grouper(conn, sql, cle=0):
    """Regroupe les lignes d'une requête en {clé: [reste des colonnes...]}."""
    out: dict = {}
    for row in conn.execute(sql):
        vals = list(row)
        out.setdefault(vals[cle], []).append(vals[:cle] + vals[cle + 1:])
    return out


def _cartes(conn) -> dict:
    """Cartes de niveau corpus, préchargées une fois (source unique pour JSON et CSV)."""
    perso_nom = {r["id"]: r["nom"] for r in conn.execute("SELECT id, nom FROM personnages")}
    return {
        "perso_nom": perso_nom,
        "locuteur": {r["region_id"]: perso_nom.get(r["personnage_id"]) for r in
                     conn.execute("SELECT region_id, personnage_id FROM bulle_locuteur")},
        "presence": {r["region_id"]: perso_nom.get(r["personnage_id"]) for r in
                     conn.execute("SELECT region_id, personnage_id FROM personnage_presence")},
        "region_attr": _grouper(conn,
            "SELECT ra.region_id, d.nom, v.valeur FROM region_attribut ra "
            "JOIN attribut_valeur v ON v.id = ra.valeur_id "
            "JOIN attribut_dimension d ON d.id = v.dimension_id ORDER BY d.nom, v.valeur"),
        "perso_attr": _grouper(conn,
            "SELECT pa.personnage_id, d.nom, v.valeur FROM personnage_attribut pa "
            "JOIN attribut_valeur v ON v.id = pa.valeur_id "
            "JOIN attribut_dimension d ON d.id = v.dimension_id ORDER BY d.nom, v.valeur"),
        "annot": {r["region_id"]: {"note": r["note"], "date_creation": r["date_creation"],
                                   "date_modification": r["date_modification"]}
                  for r in conn.execute("SELECT region_id, note, date_creation, "
                                        "date_modification FROM annotations")},
        "tags_cat": {r["label"]: {"couleur": r["couleur"], "description": r["description"]}
                     for r in conn.execute("SELECT label, couleur, description FROM tags")},
        "nb_planches": {r["album_id"]: r["n"] for r in conn.execute(
            "SELECT album_id, COUNT(*) AS n FROM planches GROUP BY album_id")},
        "ann_tags": _grouper(conn,
            "SELECT a.region_id, t.label FROM annotations a "
            "JOIN annotation_tags at ON at.annotation_id = a.id "
            "JOIN tags t ON t.id = at.tag_id ORDER BY t.label"),
        "tokens": _tokens_by_region(conn),
        "numero_editorial": _numeros_editoriaux_global(conn),
        "meta": {r["cle"]: r["valeur"] for r in conn.execute("SELECT cle, valeur FROM meta")},
        "schema_version": conn.execute("PRAGMA user_version").fetchone()[0],
    }


def _tokens_by_region(conn) -> dict:
    out: dict = {}
    for r in conn.execute("SELECT region_id, ordre, texte, lemme, pos, morph, provenance, "
                          "a_revoir, corr_auteur FROM tokens_effectifs "
                          "ORDER BY region_id, ordre"):
        out.setdefault(r["region_id"], []).append({
            "ordre": r["ordre"], "texte": r["texte"], "lemme": r["lemme"],
            "pos": r["pos"], "morph": r["morph"] or None,
            "provenance": r["provenance"], "a_revoir": r["a_revoir"],
            "auteur": r["corr_auteur"]})
    return out


def _numeros_editoriaux_global(conn) -> dict:
    out: dict = {}
    for a in conn.execute("SELECT id FROM albums"):
        out.update(database.numeros_editoriaux(conn, a["id"]))
    return out


def _paires(liste):
    return [{"dimension": dim, "valeur": val} for dim, val in (liste or [])]


# --------------------------------------------------------------------------- #
# Vue JSON — arbre imbriqué
# --------------------------------------------------------------------------- #
def collecter(conn, verbatim: bool = False) -> dict:
    c = _cartes(conn)

    vocab = []
    for d in conn.execute("SELECT id, cible, nom FROM attribut_dimension ORDER BY cible, nom"):
        vals = [r["valeur"] for r in conn.execute(
            "SELECT valeur FROM attribut_valeur WHERE dimension_id = ? ORDER BY valeur",
            (d["id"],))]
        vocab.append({"cible": d["cible"], "nom": d["nom"], "valeurs": vals})

    personnages = [{
        "id": p["id"], "nom": p["nom"], "serie": p["serie"], "notes": p["notes"],
        "attributs": _paires(c["perso_attr"].get(p["id"])),
    } for p in conn.execute("SELECT id, nom, serie, notes FROM personnages ORDER BY nom")]

    tags = [{"label": lbl, "couleur": pr["couleur"], "description": pr["description"]}
            for lbl, pr in sorted(c["tags_cat"].items())]

    def region_node(r, par_parent, cits):
        ocr = r["ocr_texte"] or ""
        node = {
            "id": r["id"], "type": r["type"],
            "x": r["x"], "y": r["y"], "w": r["w"], "h": r["h"],
            "ordre": r["ordre"], "source": r["source"], "date_creation": r["date_creation"],
            "citation": (cits.get(r["id"]) or {}).get("texte"),
            "ocr": {"present": bool(ocr.strip()), "longueur": len(ocr),
                    **({"texte": ocr} if verbatim else {})},
            "locuteur": c["locuteur"].get(r["id"]),
            "presence": c["presence"].get(r["id"]),
        }
        tg = [t[0] for t in c["ann_tags"].get(r["id"], [])]
        an = c["annot"].get(r["id"])
        if an is not None or tg:
            node["annotation"] = {
                "note": (an or {}).get("note"), "tags": tg,
                "date_creation": (an or {}).get("date_creation"),
                "date_modification": (an or {}).get("date_modification")}
        if c["region_attr"].get(r["id"]):
            node["attributs"] = _paires(c["region_attr"][r["id"]])
        if c["tokens"].get(r["id"]):
            node["tokens"] = [{k: v for k, v in t.items() if not (k == "auteur" and v is None)}
                              for t in c["tokens"][r["id"]]]
        enfants = [region_node(x, par_parent, cits) for x in par_parent.get(r["id"], [])]
        if enfants:
            node["enfants"] = enfants
        return node

    albums = []
    for a in conn.execute("SELECT * FROM albums ORDER BY id"):
        planches = []
        for p in conn.execute("SELECT * FROM planches WHERE album_id = ? ORDER BY numero, id",
                             (a["id"],)):
            regs = conn.execute("SELECT * FROM regions WHERE planche_id = ? ORDER BY ordre, id",
                               (p["id"],)).fetchall()
            cits = database.citations_regions(conn, [r["id"] for r in regs])
            par_parent: dict = {}
            for r in regs:
                par_parent.setdefault(r["parent_id"], []).append(r)
            planches.append({
                "id": p["id"], "numero": p["numero"], "role": p["role"],
                "numero_editorial": c["numero_editorial"].get(p["id"]),
                "statut": p["statut"], "validee": p["validee"],
                "verrouillee": p["verrouillee"], "date_segmentation": p["date_segmentation"],
                "largeur_px": p["largeur_px"], "hauteur_px": p["hauteur_px"],
                "chemin_tiff": p["chemin_tiff"], "chemin_web": p["chemin_web"],
                "regions": [region_node(r, par_parent, cits) for r in par_parent.get(None, [])],
            })
        albums.append({
            "id": a["id"], "titre": a["titre"], "auteur": a["auteur"], "annee": a["annee"],
            "editeur": a["editeur"], "serie": a["serie"], "description": a["description"],
            "date_import": a["date_import"], "nombre_pages": len(planches),
            "planches": planches,
        })

    return {"metadonnees_collection": {
        "genere_le": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "perimetre": {"type": "collection", "collection_id": None, "portee": "corpus entier"},
        "ocr_verbatim_inclus": verbatim,
        "paradonnee": {                       # niveau 8 — le PROGRAMME (cf. dictionnaire N8)
            "schema_version": c["schema_version"],
            "outil": version_outil(BASE_DIR),  # nom + version (déclarée) + révision git
            "environnement": environnement(),  # python + versions installées (à l'export)
            "logiciels": composants(),        # inventaire documenté : version + rôle + site + description
            "meta": c["meta"],                # modèle NLP + versions + dates de réindexation
            "a_prevoir": ["activité (run)", "journal d'événements (append-only)",
                          "activite_id par entité", "touché / date_modification",
                          "indicateurs de couverture (dérivés du journal)",
                          "licence & droits par jeu"],
        },
        "vocabulaire": vocab, "tags": tags, "personnages": personnages, "albums": albums,
    }}


# --------------------------------------------------------------------------- #
# Vue CSV — tables relationnelles (un dict {nom: (colonnes, lignes)})
# --------------------------------------------------------------------------- #
def tables(conn, verbatim: bool = False) -> dict:
    c = _cartes(conn)
    cits = database.citations_regions(
        conn, [r["id"] for r in conn.execute("SELECT id FROM regions")])
    out: dict = {}

    out["albums"] = (
        ["id", "titre", "auteur", "annee", "editeur", "serie", "description", "date_import",
         "nombre_pages"],
        [[a["id"], a["titre"], a["auteur"], a["annee"], a["editeur"], a["serie"],
          a["description"], a["date_import"], c["nb_planches"].get(a["id"], 0)]
         for a in conn.execute("SELECT * FROM albums ORDER BY id")])

    out["planches"] = (
        ["id", "album_id", "numero", "role", "numero_editorial", "statut", "validee",
         "verrouillee", "date_segmentation", "largeur_px", "hauteur_px",
         "chemin_tiff", "chemin_web"],
        [[p["id"], p["album_id"], p["numero"], p["role"],
          c["numero_editorial"].get(p["id"]), p["statut"], p["validee"], p["verrouillee"],
          p["date_segmentation"], p["largeur_px"], p["hauteur_px"],
          p["chemin_tiff"], p["chemin_web"]]
         for p in conn.execute("SELECT * FROM planches ORDER BY album_id, numero, id")])

    reg_cols = ["id", "planche_id", "album_id", "parent_id", "type", "x", "y", "w", "h",
                "ordre", "source", "date_creation", "citation", "ocr_present", "ocr_longueur"]
    if verbatim:
        reg_cols.append("ocr_texte")
    reg_cols += ["locuteur", "presence"]
    reg_rows = []
    for r in conn.execute("SELECT r.*, p.album_id AS album_id FROM regions r "
                         "JOIN planches p ON p.id = r.planche_id "
                         "ORDER BY r.planche_id, r.ordre, r.id"):
        ocr = r["ocr_texte"] or ""
        ligne = [r["id"], r["planche_id"], r["album_id"], r["parent_id"], r["type"],
                 r["x"], r["y"], r["w"], r["h"], r["ordre"], r["source"], r["date_creation"],
                 (cits.get(r["id"]) or {}).get("texte"),
                 1 if ocr.strip() else 0, len(ocr)]
        if verbatim:
            ligne.append(ocr)
        ligne += [c["locuteur"].get(r["id"]), c["presence"].get(r["id"])]
        reg_rows.append(ligne)
    out["regions"] = (reg_cols, reg_rows)

    out["tokens"] = (
        ["region_id", "ordre", "texte", "lemme", "pos", "morph", "provenance", "a_revoir",
         "auteur"],
        [[rid, t["ordre"], t["texte"], t["lemme"], t["pos"], t["morph"],
          t["provenance"], t["a_revoir"], t["auteur"]]
         for rid, toks in c["tokens"].items() for t in toks])

    out["annotations"] = (
        ["region_id", "note", "tags", "date_creation", "date_modification"],
        [[rid, (c["annot"].get(rid) or {}).get("note"),
          "|".join(t[0] for t in c["ann_tags"].get(rid, [])),
          (c["annot"].get(rid) or {}).get("date_creation"),
          (c["annot"].get(rid) or {}).get("date_modification")]
         for rid in sorted(set(c["annot"]) | set(c["ann_tags"]))])

    out["tags"] = (
        ["label", "couleur", "description"],
        [[lbl, pr["couleur"], pr["description"]] for lbl, pr in sorted(c["tags_cat"].items())])

    out["personnages"] = (
        ["id", "nom", "serie", "notes"],
        [[p["id"], p["nom"], p["serie"], p["notes"]]
         for p in conn.execute("SELECT * FROM personnages ORDER BY nom")])

    out["personnage_attributs"] = (
        ["personnage_id", "dimension", "valeur"],
        [[pid, dim, val] for pid, paires in c["perso_attr"].items()
         for dim, val in paires])

    out["region_attributs"] = (
        ["region_id", "dimension", "valeur"],
        [[rid, dim, val] for rid, paires in c["region_attr"].items()
         for dim, val in paires])

    out["vocabulaire"] = (
        ["cible", "dimension", "valeur"],
        [[d["cible"], d["nom"], v["valeur"]]
         for d in conn.execute("SELECT id, cible, nom FROM attribut_dimension "
                              "ORDER BY cible, nom")
         for v in conn.execute("SELECT valeur FROM attribut_valeur "
                              "WHERE dimension_id = ? ORDER BY valeur", (d["id"],))])

    ov = version_outil(BASE_DIR)              # provenance de l'outil (paradonnée)
    env = environnement()                     # python + versions installées (à l'export)
    out["paradonnee"] = (                     # niveau 8 — le PROGRAMME (schéma + outil + python + meta)
        ["cle", "valeur"],
        [["schema_version", c["schema_version"]],
         ["outil", ov["nom"]], ["outil_version", ov["version"]],
         ["outil_revision", ov["revision"]],
         ["python", env["python"]],
         ["ocr_verbatim_inclus", 1 if verbatim else 0]]
        + [[k, v] for k, v in c["meta"].items()])

    # Inventaire logiciel documenté (SBOM) : une ligne par composant (Python, Kumiko,
    # paquets) avec version RÉELLE + rôle dans le projet + site officiel + description
    # générique (résumé PyPI). Remplace/enrichit les anciennes lignes pkg: de paradonnée.
    out["logiciels"] = (
        ["composant", "categorie", "version", "role", "site_officiel", "description"],
        [[co["composant"], co["categorie"],
          co["version"] if co["installe"] else "(absent)",
          co["role"], co["site"], co["resume"] or ""]
         for co in composants()])
    return out


def _ecrire_table(f, cols, rows):
    w = csv.writer(f)
    w.writerow(cols)
    w.writerows(rows)


def _aplatir(obj, prefixe=""):
    """Déplie un dict/list imbriqué en lignes (clé pointée, valeur) pour la fiche."""
    lignes = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            lignes += _aplatir(v, f"{prefixe}{k}.")
    elif isinstance(obj, list):
        if all(not isinstance(x, (dict, list)) for x in obj):
            lignes.append((prefixe.rstrip("."), "; ".join(str(x) for x in obj)))
        else:
            for i, x in enumerate(obj):
                lignes += _aplatir(x, f"{prefixe}{i}.")
    else:
        lignes.append((prefixe.rstrip("."), "" if obj is None else obj))
    return lignes


_BOM = "\ufeff"          # BOM UTF-8 : permet à Excel de lire les accents
_XLSX_INJECT = ("=", "+", "-", "@")


def _neutraliser_ligne(ws, num_ligne):
    """Anti-injection : force en TEXTE toute valeur de donnée commençant par = + - @
    (sinon Excel/openpyxl l'interprète en formule). Ne touche pas aux formules qu'on
    écrit sciemment (les hyperliens de l'onglet `arbre`)."""
    for cell in ws[num_ligne]:
        if isinstance(cell.value, str) and cell.value[:1] in _XLSX_INJECT:
            cell.data_type = "s"


def _ecrire_xlsx(tbls: dict, arbre: dict, fiche: dict, chemin: str) -> None:
    """Classeur XLSX. Onglets : `fiche` (roll-up descriptif) · `arbre` (hiérarchie
    repliable + liens vers le détail) · `_tables` (index) · une table par niveau.
    En-tête gelé/gras, filtres. `openpyxl` requis (import protégé : hors noyau —
    JSON/CSV marchent sans)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("XLSX demandé mais 'openpyxl' est absent "
                         "(pip install -r requirements-export.txt).")
    gras = Font(bold=True)
    lien = Font(color="1155CC", underline="single")
    wb = Workbook()
    wb.remove(wb.active)

    # --- fiche descriptive (roll-up : couverture / provenance / droits) ------ #
    fs = wb.create_sheet("fiche")
    fs.append(["champ", "valeur"])
    for cle, val in _aplatir(fiche):
        fs.append([cle, val])
        _neutraliser_ligne(fs, fs.max_row)
    for c in fs[1]:
        c.font = gras
    fs.freeze_panes = "A2"
    fs.column_dimensions["A"].width = 40
    fs.column_dimensions["B"].width = 64

    # --- index --------------------------------------------------------------- #
    idx = wb.create_sheet("_tables")
    idx.append(["table", "lignes", "colonnes"])
    for nom, (cols, rows) in tbls.items():
        idx.append([nom, len(rows), ", ".join(cols)])
    for c in idx[1]:
        c.font = gras
    idx.freeze_panes = "A2"

    # --- tables détail (on retient la ligne de chaque entité pour les liens) - #
    pos: dict = {}
    detail = []
    for nom, (cols, rows) in tbls.items():
        ws = wb.create_sheet(title=nom[:31])
        ws.append(cols)
        for i, r in enumerate(rows, start=2):
            ws.append(list(r))
            _neutraliser_ligne(ws, i)
            pos[(nom, r[0])] = i          # r[0] = id (albums/planches/regions/personnages)
        for c in ws[1]:
            c.font = gras
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
        detail.append(ws)

    # --- arbre : hiérarchie repliable + boîtes (px master) + hyperliens ------ #
    arb = wb.create_sheet("arbre")
    arb.append(["hiérarchie", "type", "x", "y", "w", "h",
                "citation", "locuteur / présence", "détail"])
    for c in arb[1]:
        c.font = gras
    arb.sheet_properties.outlinePr.summaryBelow = False   # le parent est AU-DESSUS
    arb.freeze_panes = "A2"
    ln = [1]

    def ajoute(libelle, typ, box, citation, qui, table, ident, niveau):
        ln[0] += 1
        r = ln[0]
        arb.cell(r, 1, libelle)
        arb.cell(r, 2, typ)
        x, y, w, h = box if box else ("", "", "", "")
        arb.cell(r, 3, x)
        arb.cell(r, 4, y)
        arb.cell(r, 5, w)
        arb.cell(r, 6, h)
        arb.cell(r, 7, citation or "")
        arb.cell(r, 8, qui or "")
        for col in range(1, 9):           # anti-injection (le lien col. 9 reste une formule)
            cell = arb.cell(r, col)
            if isinstance(cell.value, str) and cell.value[:1] in _XLSX_INJECT:
                cell.data_type = "s"
        p = pos.get((table, ident))
        if p:
            arb.cell(r, 9, f'=HYPERLINK("#{table}!A{p}","voir")').font = lien
        arb.row_dimensions[r].outline_level = min(niveau, 7)

    def walk(reg, niveau):
        qui = reg.get("locuteur") or reg.get("presence") or ""
        ajoute("  " * niveau + reg["type"], reg["type"],
               (reg["x"], reg["y"], reg["w"], reg["h"]),
               reg.get("citation"), qui, "regions", reg["id"], niveau)
        for e in reg.get("enfants", []):
            walk(e, niveau + 1)

    for a in arbre["metadonnees_collection"]["albums"]:
        ajoute(f"Album — {a['titre']}", "album", None, "", a.get("annee") or "",
               "albums", a["id"], 0)
        for p in a["planches"]:
            ed = p["numero_editorial"]
            lbl = f"  Planche {p['numero']}" + (f" · pl.{ed}" if ed else " · paratexte")
            ajoute(lbl, "planche", (0, 0, p["largeur_px"], p["hauteur_px"]),
                   "", p["statut"], "planches", p["id"], 1)
            for reg in p["regions"]:
                walk(reg, 2)
    for col, w in (("A", 42), ("G", 18), ("H", 22)):
        arb.column_dimensions[col].width = w

    # --- ordre des onglets : fiche, arbre, index, puis les tables ------------ #
    wb._sheets = [fs, arb, idx] + detail
    wb.save(chemin)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def _connexion_ro():
    conn = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Export des métadonnées réelles d'une "
                                             "collection (corpus entier).")
    ap.add_argument("--json", nargs="?", const="-", metavar="FICHIER",
                    help="arbre JSON (défaut si aucune sortie demandée ; '-' = stdout)")
    ap.add_argument("--csv-dir", metavar="DOSSIER",
                    help="écrit une table CSV par niveau dans ce dossier")
    ap.add_argument("--zip", metavar="FICHIER",
                    help="écrit toutes les tables CSV dans une archive zip")
    ap.add_argument("--xlsx", metavar="FICHIER",
                    help="écrit un classeur XLSX (un onglet par table ; requiert openpyxl)")
    ap.add_argument("--verbatim", action="store_true",
                    help="inclut le texte OCR (contenu restreint)")
    args = ap.parse_args(argv)

    veut_tables = bool(args.csv_dir or args.zip or args.xlsx)
    if args.json is None and not veut_tables:
        args.json = "-"

    with _connexion_ro() as conn:
        besoin_arbre = args.json is not None or bool(args.xlsx)
        arbre = collecter(conn, verbatim=args.verbatim) if besoin_arbre else None
        tbls = tables(conn, verbatim=args.verbatim) if veut_tables else None
        fiche = None
        if args.xlsx:
            import description_collection
            fiche = description_collection.collecter(conn)[0]["description_collection"]

    if args.json is not None:
        texte = json.dumps(arbre, ensure_ascii=False, indent=2)
        if args.json == "-":
            print(texte)
        else:
            with open(args.json, "w", encoding="utf-8") as f:
                f.write(texte + "\n")
            print(f"JSON écrit : {args.json}", file=sys.stderr)

    if args.csv_dir:
        os.makedirs(args.csv_dir, exist_ok=True)
        for nom, (cols, rows) in tbls.items():
            with open(os.path.join(args.csv_dir, f"{nom}.csv"), "w",
                      encoding="utf-8-sig", newline="") as f:   # BOM → accents OK dans Excel
                _ecrire_table(f, cols, rows)
            print(f"  {nom}.csv : {len(rows)} lignes", file=sys.stderr)
        print(f"Tables CSV écrites : {args.csv_dir}", file=sys.stderr)

    if args.zip:
        with zipfile.ZipFile(args.zip, "w", zipfile.ZIP_DEFLATED) as z:
            for nom, (cols, rows) in tbls.items():
                buf = io.StringIO()
                _ecrire_table(buf, cols, rows)
                z.writestr(f"{nom}.csv", _BOM + buf.getvalue())   # BOM UTF-8 (Excel)
        print(f"Archive écrite : {args.zip}", file=sys.stderr)

    if args.xlsx:
        _ecrire_xlsx(tbls, arbre, fiche, args.xlsx)
        print(f"Classeur XLSX écrit : {args.xlsx}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
